import openpyxl

workbook=openpyxl.load_workbook("data.xlsx")

sheet=workbook.active
print(sheet.cell(row=1,column=2).value)
print(sheet['B3'].value)
print(sheet.max_row)
print(sheet.max_column)

sheet['B4'].value="Sanjay"
print(sheet['B4'].value)

dict={}
for i in range(2,sheet.max_row+1):
    name=sheet.cell(row=i, column=1).value
    if  name== "Testcase2":
        for j in range(2,sheet.max_column+1):
            # print(sheet.cell(row=i,column=j).value,end=" ")
            dict[sheet.cell(1,j).value]=sheet.cell(row=i,column=j).value
print(dict)