import openpyxl


class HomePageData:

    @staticmethod
    def getDataSet(testcase_name):
        dict = {}
        workbook = openpyxl.load_workbook("../data.xlsx")
        sheet = workbook.active
        for i in range(2, sheet.max_row + 1):
            name = sheet.cell(row=i, column=1).value
            if name == testcase_name:
                for j in range(2, sheet.max_column + 1):
                    # print(sheet.cell(row=i,column=j).value,end=" ")
                    dict[sheet.cell(1, j).value] = sheet.cell(row=i, column=j).value
        return [dict]