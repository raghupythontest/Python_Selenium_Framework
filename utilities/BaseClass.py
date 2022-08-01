import inspect
import logging

import openpyxl
import pytest
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait


@pytest.mark.usefixtures("setup")
class BaseClass:
    def wait_until_element_presence(self,locator):
        wait = WebDriverWait(self.driver, 10)
        wait.until(expected_conditions.visibility_of_element_located(locator))

    def selectByVisibleText(self,locator,text):
        gender=Select(locator)
        gender.select_by_visible_text(text)

    def getLogger(self):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        # Creting a fileHandler
        fh = logging.FileHandler("logfile.log")
        # Defining format of the log
        formatter = logging.Formatter("%(asctime)s :%(levelname)s :%(name)s :%(message)s")
        # setting format to filehandler
        fh.setFormatter(formatter)
        # Now file handler has a knowlege about where to print and which format it needs to print,add the file handler to logger object
        logger.addHandler(fh)
        # setting level if we set the level to warning, it will look print only warning,error,critical(basically it follows hierarchy)
        logger.setLevel(logging.DEBUG)
        return logger

    @staticmethod
    def getDataSet(testcase_name):
        dict = {}
        workbook = openpyxl.load_workbook("data.xlsx")
        sheet = workbook.active
        for i in range(2, sheet.max_row + 1):
            name = sheet.cell(row=i, column=1).value
            if name ==testcase_name:
                for j in range(2, sheet.max_column + 1):
                    # print(sheet.cell(row=i,column=j).value,end=" ")
                    dict[sheet.cell(1, j).value] = sheet.cell(row=i, column=j).value
        return [dict]
